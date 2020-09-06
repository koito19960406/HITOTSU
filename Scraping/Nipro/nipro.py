#!/usr/bin/env python
# coding: utf-8

# # Import Libraries

# In[6]:


from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import datetime
import csv
import openpyxl
from openpyxl import Workbook
import re


# # Create Functions

# ## Scrape info

# In[9]:


class Nipro:
    # prepare the input date
    def __init__(self, year, month, day):
        input_year=str(year)
        input_month=str(month)
        input_day=str(day)
        self.input_date=input_year+'年'+input_month+'月'+input_day+'日'

        # modify month and day for output_date
        output_year=str(year)
        if month<10:
            output_month='0'+str(month)
        else:
            output_month=str(month)
        if day<10:
            output_day='0'+str(day)
        else:
            output_day=str(day)
        self.output_date=int(output_year+output_month+output_day)



    # scrape the info
    def scrape(self):
        #open chrome in incognito mode
        options = webdriver.ChromeOptions()
        options.add_argument(' -- incognito')
        browser = webdriver.Chrome(chrome_options=options)

        # deal with the first "medical staff?" question
        browser.get('http://med.nipro.co.jp/index')
        timeout = 10
        try:
            WebDriverWait(browser, timeout).until(
            EC.visibility_of_element_located(
            (By.XPATH, '//*[@id="j_id0:j_id16"]/ul/li[1]/a')
            )
            )
        except TimeoutException:
            print('Timed Out Waiting for page to load')
            browser.quit()
        login_btn=browser.find_element_by_xpath('//*[@id="j_id0:j_id16"]/ul/li[1]/a')
        login_btn.click()
        browser.implicitly_wait(3)

        # activate medical news button
        activate_btn=browser.find_element_by_xpath('/html/body/div[1]/div[4]/div[1]/div[1]/div[2]/ul/li[2]/span')
        browser.execute_script('arguments[0].click();', activate_btn)

        # Get info
        # Go to the list
        date_list=browser.find_element_by_xpath('//*[@id="j_id0:j_id83"]/dl').find_elements_by_css_selector('dt')
        news_list=browser.find_element_by_xpath('//*[@id="j_id0:j_id83"]/dl').find_elements_by_css_selector('dd')
    
        # Go through the list
        result=[]
        date_count=0
        for dates in date_list:
            # Get URL and title if date == input date
            dates_text=dates.text
            date = re.search('\d{4}年\d{1,}月\d{1,}日', dates_text).group()
            if date==self.input_date:

                # Get link and title 
                news_count=0
                for news in news_list:
                    if date_count==news_count:
                        news_url=news.find_element_by_css_selector('a').get_attribute('href')
                        news_title=news.find_element_by_tag_name('a').text
                        news_category=dates.find_element_by_css_selector('span.ph_tag.category_another').text
                        # If the title contains "新発売", them return 1 as new_product
                        new_product_condition_1='新発売'
                        if new_product_condition_1 in news_title:
                            new_product=1
                        else:
                            new_product=0
                        # Append the info to the list
                        result.append([self.output_date,news_category,news_title,news_url,new_product])
                        break
                    else:
                        news_count+=1
                date_count+=1

        # close the browser        
        browser.quit()
        return result

    # store info into csv
    def to_csv(self):
        # get result
        result = self.scrape()
        result_len = len(result)

        # check if the result is empty
        if result_len == 0:
            return
        
        # get date for checking their existence later
        date=str(result[0][0])

        # get row number
        # try to open the csv file
        try:
            with open('Nipro.csv') as csvfile:
                reader = csv.reader(csvfile)
                # check if the title we are trying to add is already there    
                for row in reader:
                    # the date is already there, dont add anything
                    if row[0]==date:
                        return print('Already added to csv')
        # if there's no such file, create a new file 
        except FileNotFoundError:
            with open('Nipro.csv','w') as csvfile:
                pass
        
        # add new data
        with open('Nipro.csv', 'a') as csvfile:
            writer = csv.writer(csvfile)
            for i in range(result_len):
                writer.writerow([result[i][0], 
                                '血液浄化', 
                                '医療機器管理', 
                                '',
                                '',
                                'ニプロ', 
                                result[i][1],
                                result[i][2], 
                                result[i][3], 
                                result[i][4]])

    # store info into excel
    def to_excel(self):
        # get result
        result = self.scrape()
        result_len = len(result)

        # check if the result is empty
        result_len = len(result)
        if result_len == 0:
            return

        # get date for checking their existence later
        date=result[0][0]

        # try to open the workbook
        try:
            wb = openpyxl.load_workbook('Nipro.xlsx')
            ws = wb['Sheet1']
            for row in ws.iter_rows(values_only=True):
                # the date is already there, dont add anything
                if row[0]==date:
                    return print('Already added to excel')
        # if we cannot open it, we create a new one
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.create_sheet('Sheet1')
            ws.append(['日付',
                    'カテゴリコード１',
                    'カテゴリコード２',
                    'カテゴリコード３',
                    'メーカーコード',
                    'メーカー名称',
                    '新着記事カテゴリ',
                    '新着記事タイトル',
                    '新着記事URL',
                    '新製品記事'])


        # check the last row in excel
        last_row = ws.max_row

        # update excel
        # can handle up to 3 news on the same day
        for i in range(result_len):
            ws.cell(row = last_row + i + 1, column = 1, value = result[i][0]) # 日付
            ws.cell(row = last_row + i + 1, column = 2, value = '血液浄化') # カテゴリコード１
            ws.cell(row = last_row + i + 1, column = 3, value = '医療機器管理') # カテゴリコード2
            ws.cell(row = last_row + i + 1, column = 4, value = '') # カテゴリコード3

            # add メーカーコード

            ws.cell(row = last_row + i + 1, column = 6, value = 'ニプロ') # メーカー名称
            ws.cell(row = last_row + i + 1, column = 7, value = result[i][1]) # 新着記事カテゴリ
            ws.cell(row = last_row + i + 1, column = 8, value = result[i][2]) # 新着記事タイトル
            ws.cell(row = last_row + i + 1, column = 9, value = result[i][3]) # 新着記事URL
            ws.cell(row = last_row + i + 1, column = 10, value = result[i][4]) # 新製品記事

        wb.save('Nipro.xlsx')


# # Run the Function

# In[10]:


if __name__=='__main__':
    # set the beginning and the end of the dates to get info
    start_date = datetime.date(2020, 8, 1)
    end_date = datetime.date(2020, 8, 31)
    # interval is the interval between each loop (e.g., in this case, it's 1 day)
    interval = datetime.timedelta(days=1)

    # run the loop
    while start_date <= end_date:
        year=start_date.year
        month=start_date.month
        day=start_date.day
        nipro=Nipro(year,month,day)
        nipro.to_csv()
        nipro.to_excel()
        start_date += interval


# In[ ]:




