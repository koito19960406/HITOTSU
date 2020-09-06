#!/usr/bin/env python
# coding: utf-8

# # Import Libraries

# In[10]:


from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import datetime
import csv
import openpyxl
from openpyxl import Workbook


# # Create Functions

# In[11]:


class Nihonkoden:
    # prepare the input date
    def __init__(self, year, month, day):
        year=str(year)
        if month<10:
            month='0'+str(month)
        else:
            month=str(month)
        if day<10:
            day='0'+str(day)
        else:
            day=str(day)
        self.input_date=year+'年'+month+'月'+day+'日'
        self.output_date=int(year+month+day)
    
    # scrape the info
    def scrape(self):
        #open chrome in incognito mode
        options = webdriver.ChromeOptions()
        options.add_argument(' -- incognito')
        browser = webdriver.Chrome(chrome_options=options)

        # deal with the first "medical staff?" question
        browser.get('https://www.nihonkohden.co.jp/iryo/index.html?URL=/iryo/index2.html')
        timeout = 10
        try:
            WebDriverWait(browser, timeout).until(
            EC.visibility_of_element_located(
            (By.XPATH, '//*[@id="pagetop"]/div[4]/div/div/div/div/div/div/div/div[2]/a[1]')
            )
            )
        except TimeoutException:
            print('Timed Out Waiting for page to load')
            browser.quit()

        # Click the yes button
        login_btn=browser.find_element_by_xpath('//*[@id="pagetop"]/div[4]/div/div/div/div/div/div/div/div[2]/a[1]')
        login_btn.click()
        browser.implicitly_wait(3)

        # Get info
        # Go to production info section
        news_list=browser.find_element_by_class_name('infoset').find_elements_by_css_selector('tr')
        # Go through the list
        result=[]
        for news in news_list:
            date=news.find_element_by_css_selector('th').text
            if date==self.input_date:
                news_url=news.find_element_by_css_selector('a').get_attribute('href')
                news_title=news.find_element_by_css_selector('a').text
                # If the title contains "新商品", them return 1 as new_product
                new_product_condition_1='新製品'
                if new_product_condition_1 in news_title:
                    new_product=1
                else:
                    new_product=0
                result.append([self.output_date,news_title,news_url,new_product])

        # close the browser        
        browser.quit()
        return result

    # store info into csv
    def to_csv(self):
        # get result
        result = self.scrape()
        result_len = len(result)

        # check if the result is empty
        if result_len == 0:
            return
        
        # get date for checking their existence later
        date=str(result[0][0])

        # get row number
        # try to open the csv file
        try:
            with open('Nihonkoden.csv') as csvfile:
                reader = csv.reader(csvfile)
                # check if the title we are trying to add is already there    
                for row in reader:
                    # the date is already there, dont add anything
                    if row[0]==date:
                        return print('Already added to csv')
        # if there's no such file, create a new file 
        except FileNotFoundError:
            with open('Nihonkoden.csv','w') as csvfile:
                pass
        
        # add new data
        with open('Nihonkoden.csv', 'a') as csvfile:
            writer = csv.writer(csvfile)
            for i in range(result_len):
                writer.writerow([result[i][0], 
                                '手術室', 
                                '医療機器管理', 
                                '人工呼吸器',
                                '',
                                '日本光電工業', 
                                '',
                                result[i][1], 
                                result[i][2], 
                                result[i][3]])

    # store info into excel
    def to_excel(self):
        # get result
        result = self.scrape()
        result_len = len(result)

        # check if the result is empty
        result_len = len(result)
        if result_len == 0:
            return

        # get date for checking their existence later
        date=result[0][0]

        # try to open the workbook
        try:
            wb = openpyxl.load_workbook('Nihonkoden.xlsx')
            ws = wb['Sheet1']
            for row in ws.iter_rows(values_only=True):
                # the date is already there, dont add anything
                if row[0]==date:
                    return print('Already added to excel')
        # if we cannot open it, we create a new one
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.create_sheet('Sheet1')
            ws.append(['日付',
                    'カテゴリコード１',
                    'カテゴリコード２',
                    'カテゴリコード３',
                    'メーカーコード',
                    'メーカー名称',
                    '新着記事カテゴリ',
                    '新着記事タイトル',
                    '新着記事URL',
                    '新製品記事'])


        # check the last row in excel
        last_row = ws.max_row

        # update excel
        # can handle up to 3 news on the same day
        for i in range(result_len):
            ws.cell(row = last_row + i + 1, column = 1, value = result[i][0]) # 日付
            ws.cell(row = last_row + i + 1, column = 2, value = '手術室') # カテゴリコード１
            ws.cell(row = last_row + i + 1, column = 3, value = '医療機器管理') # カテゴリコード2
            ws.cell(row = last_row + i + 1, column = 4, value = '人工呼吸器') # カテゴリコード3

            # add メーカーコード

            ws.cell(row = last_row + i + 1, column = 6, value = '日本光電工業') # メーカー名称

            # add 新着記事カテゴリ
            
            ws.cell(row = last_row + i + 1, column = 8, value = result[i][1]) # 新着記事タイトル
            ws.cell(row = last_row + i + 1, column = 9, value = result[i][2]) # 新着記事URL
            ws.cell(row = last_row + i + 1, column = 10, value = result[i][3]) # 新製品記事

        wb.save('Nihonkoden.xlsx')


# # Run the Function

# In[12]:


if __name__=='__main__':
    # set the beginning and the end of the dates to get info
    start_date = datetime.date(2020, 8, 1)
    end_date = datetime.date(2020, 8, 31)
    # interval is the interval between each loop (e.g., in this case, it's 1 day)
    interval = datetime.timedelta(days=1)

    # run the loop
    while start_date <= end_date:
        year=start_date.year
        month=start_date.month
        day=start_date.day
        nihonkoden=Nihonkoden(year,month,day)
        nihonkoden.to_csv()
        nihonkoden.to_excel()
        start_date += interval


# In[ ]:




