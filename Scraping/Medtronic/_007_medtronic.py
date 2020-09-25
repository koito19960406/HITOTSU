#!/usr/bin/env python
# coding: utf-8

# # Import Libraries

# In[12]:


from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import datetime
import csv
import openpyxl
from openpyxl import Workbook


# # Create Functions

# In[13]:


class Medtronic:
    # prepare the input date
    def __init__(self, year, month, day):
        input_year=str(year)
        input_month=str(month)
        input_day=str(day)
        self.input_date=input_year+'/'+input_month+'/'+input_day

        # modify month and day for output_date
        output_year=str(year)
        if month<10:
            output_month='0'+str(month)
        else:
            output_month=str(month)
        if day<10:
            output_day='0'+str(day)
        else:
            output_day=str(day)
        self.output_date=int(output_year+output_month+output_day)
    
    # scrape the info
    def scrape(self):
        #open chrome in incognito mode
        options = webdriver.ChromeOptions()
        options.add_argument(' -- incognito')
        browser = webdriver.Chrome(chrome_options=options)

        # Go to the website
        browser.get('https://www.medtronic.com/jp-ja/about/news.html')

        # wait for 10 seconds to load the website
        timeout = 10
        try:
            WebDriverWait(browser, timeout).until(
            EC.visibility_of_element_located(
            (By.XPATH, '/html/body/div[2]/div[5]/div[3]/section/div[2]/div/ul')
            )
            )
        except TimeoutException:
            print('Timed Out Waiting for page to load')
            browser.quit()

        # Get info
        # Go to the list
        news_list=browser.find_element_by_xpath('/html/body/div[2]/div[5]/div[3]/section/div[2]/div/ul').find_elements_by_css_selector('li')
        # Go through the list
        result=[]
        for news in news_list:
            date = news.find_element_by_css_selector('p>b').text
            # Get URL and title if date == input date
            if date==self.input_date:
                # Get link and title
                news_url=news.find_element_by_css_selector('p>a').get_attribute('href')
                news_title=news.find_element_by_tag_name('p').text.split('\n')[1]
                # If the title contains "販売"&"開始", them return 1 as new_product
                new_product_condition_1='販売'
                new_product_condition_2='開始'
                if (new_product_condition_1 in news_title)&(new_product_condition_2 in news_title):
                    new_product=1
                else:
                    new_product=0
                # Append the info to the list
                result.append([self.output_date,news_title,news_url,new_product])
                
        # close the browser
        browser.quit()            
        return result

    # store info into csv
    def to_csv(self):
        # get result
        result = self.scrape()
        result_len = len(result)

        # check if the result is empty
        if result_len == 0:
            return
        
        # get date for checking their existence later
        date=str(result[0][0])

        # get row number
        # try to open the csv file
        try:
            with open('product_info.csv') as csvfile:
                reader = csv.reader(csvfile)
                # check if the title we are trying to add is already there    
                for row in reader:
                    # the date is already there, dont add anything
                    if (row[0]==date) and (row[4]==7):
                        return print('Already added to csv')
        # if there's no such file, create a new file 
        except FileNotFoundError:
            with open('product_info.csv','w') as csvfile:
                pass
        
        # add new data
        with open('product_info.csv', 'a') as csvfile:
            writer = csv.writer(csvfile)
            for i in range(result_len):
                writer.writerow([result[i][0], 
                                '心臓カテーテル', 
                                '不整脈', 
                                '手術室',
                                7,
                                '日本メドトロニック', 
                                '',
                                result[i][1], 
                                result[i][2], 
                                result[i][3]])

    # store info into excel
    def to_excel(self):
        # get result
        result = self.scrape()
        result_len = len(result)

        # check if the result is empty
        result_len = len(result)
        if result_len == 0:
            return

        # get date for checking their existence later
        date=result[0][0]

        # try to open the workbook
        try:
            wb = openpyxl.load_workbook('product_info.xlsx')
            ws = wb['Sheet1']
            for row in ws.iter_rows(values_only=True):
                # the date is already there, dont add anything
                if (row[0]==date) and (row[4]==7):
                    return print('Already added to excel')
        # if we cannot open it, we create a new one
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.create_sheet('Sheet1')
            ws.append(['日付',
                    'カテゴリコード１',
                    'カテゴリコード２',
                    'カテゴリコード３',
                    'メーカーコード',
                    'メーカー名称',
                    '新着記事カテゴリ',
                    '新着記事タイトル',
                    '新着記事URL',
                    '新製品記事'])


        # check the last row in excel
        last_row = ws.max_row

        # update excel
        # can handle up to 3 news on the same day
        for i in range(result_len):
            ws.cell(row = last_row + i + 1, column = 1, value = result[i][0]) # 日付
            ws.cell(row = last_row + i + 1, column = 2, value = '心臓カテーテル') # カテゴリコード１
            ws.cell(row = last_row + i + 1, column = 3, value = '不整脈') # カテゴリコード2
            ws.cell(row = last_row + i + 1, column = 4, value = '手術室') # カテゴリコード3
            ws.cell(row = last_row + i + 1, column = 5, value = 7) # メーカーコード
            ws.cell(row = last_row + i + 1, column = 6, value = '日本メドトロニック') # メーカー名称

            # add 新着記事カテゴリ
            
            ws.cell(row = last_row + i + 1, column = 8, value = result[i][1]) # 新着記事タイトル
            ws.cell(row = last_row + i + 1, column = 9, value = result[i][2]) # 新着記事URL
            ws.cell(row = last_row + i + 1, column = 10, value = result[i][3]) # 新製品記事

        wb.save('product_info.xlsx')


# # Run the Functions

# In[14]:


if __name__=='__main__':
    # set the beginning and the end of the dates to get info
    start_date = datetime.date(2020, 8, 1)
    end_date = datetime.date(2020, 8, 31)
    # interval is the interval between each loop (e.g., in this case, it's 1 day)
    interval = datetime.timedelta(days=1)

    # run the loop
    while start_date <= end_date:
        year=start_date.year
        month=start_date.month
        day=start_date.day
        medtronic=Medtronic(year,month,day)
        medtronic.to_csv()
        medtronic.to_excel()
        start_date += interval


# In[ ]:




